#!/usr/bin/env python3
"""Run Workflow D against a Dependency-Track findings JSON file.

This is the thin normalizer that converts DT findings -> the unified CVE event
schema, then POSTs the batch to a running Workflow D service. Workflow B
remains untouched.

Examples:
    # Against a local running service:
    python scripts/analyze_dt_findings.py \\
        --vulns /Users/me/Downloads/please/vulns.json \\
        --repo-root /Users/me/work/nios \\
        --api http://127.0.0.1:8088 \\
        --severities CRITICAL HIGH \\
        --limit 20

    # In-process (no HTTP):
    python scripts/analyze_dt_findings.py \\
        --vulns vulns.json --repo-root /path/to/nios --in-process
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


_CVE_RE = re.compile(r"(?:UBUNTU-)?(CVE-\d{4}-\d+)", re.IGNORECASE)


def _extract_cve_id(vuln_id: str) -> str:
    m = _CVE_RE.search(vuln_id or "")
    return m.group(1).upper() if m else (vuln_id or "UNKNOWN")


def _ecosystem_from_purl(purl: str | None) -> str | None:
    if not purl or not purl.startswith("pkg:"):
        return None
    try:
        return purl.split(":", 1)[1].split("/", 1)[0]
    except (IndexError, ValueError):
        return None


def to_event(finding: dict, build_context: dict) -> dict:
    vuln = finding.get("vulnerability") or {}
    comp = finding.get("component") or {}
    severity = (vuln.get("severity") or "UNASSIGNED").upper()

    def _f(x):
        try:
            return float(x) if x not in (None, "") else None
        except (TypeError, ValueError):
            return None

    cvss = _f(vuln.get("cvssV3")) or _f(vuln.get("cvssV2"))
    epss = _f(vuln.get("epssScore"))
    published = vuln.get("published")
    if isinstance(published, str) and not published.strip():
        published = None
    purl = comp.get("purl") or None
    description = vuln.get("description")
    if isinstance(description, str) and not description.strip():
        description = None
    return {
        "cve_id": _extract_cve_id(vuln.get("vulnId", "")),
        "severity": {
            "label": severity,
            "cvss": cvss,
            "epss": epss,
            "exploited": False,
        },
        "component": {
            "name": comp.get("name", ""),
            "current_version": str(comp.get("version", "")) or None,
            "ecosystem": _ecosystem_from_purl(purl),
            "purl": purl,
            "fixed_version": None,  # DT findings don't carry it; advisory enrichment later
            "group": comp.get("group") or None,
        },
        "source": "dependency_track",
        "branches_affected": [],
        "filed_at": published,
        "external_links": [],
        "description": description,
        "references": vuln.get("references"),
        "build_context": build_context,
    }


def load_findings(path: Path) -> tuple[list[dict], dict]:
    """Load findings from a DependencyTrack export.

    Supports three formats:
      * .json  — native DT findings export (`{"findings": [...], "project": {...}}`)
      * .xlsx  — DT spreadsheet export with one row per vuln (columns:
                 CVE_ID, Severity, CVSS_v3_Score, CVSS_v2_Score, EPSS_Score,
                 EPSS_Percentile, Component_Name, Component_Version,
                 Component_Group, CWE_IDs, Published, Source, Description)
      * .csv   — same columns as the xlsx form
    """
    suffix = path.suffix.lower()
    if suffix == ".json":
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        findings = data.get("findings") or []
        project = data.get("project") or {}
        build_context = {
            "project_name": project.get("name"),
            "project_version": project.get("version"),
            "project_uuid": project.get("uuid"),
        }
        return findings, build_context

    if suffix in {".xlsx", ".xlsm", ".csv"}:
        return _load_findings_tabular(path)

    raise ValueError(f"unsupported vulns file type: {suffix} (expected .json/.xlsx/.csv)")


def _load_findings_tabular(path: Path) -> tuple[list[dict], dict]:
    """Read rows from an xlsx or csv DT export and normalize to the same
    shape `to_event` consumes.
    """
    rows: list[dict] = []
    suffix = path.suffix.lower()

    if suffix == ".csv":
        import csv
        with path.open("r", encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "openpyxl is required to read .xlsx vulns files: pip install openpyxl"
            ) from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            rec = {h: ("" if v is None else v) for h, v in zip(headers, row) if h}
            rows.append(rec)

    def _g(rec: dict, *keys: str) -> str:
        """Case-insensitive column getter that tolerates spaces/underscores."""
        norm = {re.sub(r"[\s_]+", "", k).lower(): v for k, v in rec.items()}
        for k in keys:
            v = norm.get(re.sub(r"[\s_]+", "", k).lower())
            if v not in (None, ""):
                return str(v).strip()
        return ""

    findings: list[dict] = []
    for rec in rows:
        cve_id = _g(rec, "CVE_ID", "CVE-ID", "vulnId", "vulnerability")
        if not cve_id:
            continue
        sev = _g(rec, "Severity") or "UNASSIGNED"
        cvss3 = _g(rec, "CVSS_v3_Score", "cvssV3", "CVSSv3")
        cvss2 = _g(rec, "CVSS_v2_Score", "cvssV2", "CVSSv2")
        epss = _g(rec, "EPSS_Score", "epss")
        epss_pct = _g(rec, "EPSS_Percentile", "epssPercentile")
        cwe_raw = _g(rec, "CWE_IDs", "CWE")
        cwes = [c.strip() for c in re.split(r"[,;|]", cwe_raw) if c.strip()] if cwe_raw else []

        finding = {
            "vulnerability": {
                "vulnId": cve_id,
                "severity": sev,
                "description": _g(rec, "Description"),
                "source": _g(rec, "Source") or "NVD",
                "published": _g(rec, "Published"),
                "cwes": cwes,
                "cvssV3": cvss3,
                "cvssV2": cvss2,
                "epssScore": epss,
                "epssPercentile": epss_pct,
                "references": None,
            },
            "component": {
                "name": _g(rec, "Component_Name", "ComponentName"),
                "version": _g(rec, "Component_Version", "ComponentVersion"),
                "group": _g(rec, "Component_Group", "ComponentGroup"),
                "purl": _g(rec, "purl", "PURL"),
            },
        }
        findings.append(finding)

    build_context = {
        "project_name": path.stem,
        "project_version": None,
        "project_uuid": None,
        "source_file": str(path),
        "source_format": suffix.lstrip("."),
    }
    return findings, build_context


def filter_findings(
    findings: list[dict],
    severities: list[str] | None,
    components: list[str] | None,
    limit: int | None,
) -> list[dict]:
    sev_set = {s.upper() for s in severities} if severities else None
    comp_set = {c.lower() for c in components} if components else None
    out: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    for f in findings:
        vuln = f.get("vulnerability") or {}
        comp = f.get("component") or {}
        sev = (vuln.get("severity") or "").upper()
        name = (comp.get("name") or "").lower()
        if sev_set and sev not in sev_set:
            continue
        if comp_set and name not in comp_set:
            continue
        key = (
            vuln.get("vulnId", ""),
            comp.get("name", ""),
            str(comp.get("version", "")),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(f)
        if limit and len(out) >= limit:
            break
    return out


def batch(events: list[dict], size: int) -> list[list[dict]]:
    return [events[i : i + size] for i in range(0, len(events), size)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize DT findings and run Workflow D")
    parser.add_argument("--vulns", required=True, help="Path to vulns.json from Workflow B")
    parser.add_argument("--repo-root", required=True, help="Path to the source repo (NIOS)")
    parser.add_argument("--api", default="http://127.0.0.1:8088",
                        help="Base URL of a running Workflow D service")
    parser.add_argument("--in-process", action="store_true",
                        help="Skip HTTP and call the orchestrator in-process")
    parser.add_argument("--mode", default="standard",
                        choices=["standard", "urgent", "ad_hoc"])
    parser.add_argument("--severities", nargs="*", default=None,
                        help="Filter by severity, e.g. CRITICAL HIGH")
    parser.add_argument("--components", nargs="*", default=None,
                        help="Filter by component name (case-insensitive)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Cap the number of CVEs to analyze")
    parser.add_argument("--batch-size", type=int, default=10,
                        help="CVEs per /analyze request")
    parser.add_argument("--out", default=None,
                        help="Optional path to write the normalized event JSON")
    args = parser.parse_args()

    vulns_path = Path(args.vulns).resolve()
    repo_root = Path(args.repo_root).resolve()
    if not vulns_path.exists():
        sys.exit(f"vulns file not found: {vulns_path}")
    if not repo_root.exists():
        sys.exit(f"repo root not found: {repo_root}")

    findings, build_context = load_findings(vulns_path)
    print(f"Loaded {len(findings)} findings from {vulns_path}")

    selected = filter_findings(findings, args.severities, args.components, args.limit)
    print(f"Selected {len(selected)} findings after filters")

    build_context["repo_root"] = str(repo_root)
    events = [to_event(f, build_context) for f in selected]

    if args.out:
        Path(args.out).write_text(
            json.dumps({"mode": args.mode, "cves": events}, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"Wrote normalized batch -> {args.out}")

    if not events:
        print("Nothing to analyze.")
        return

    batches = batch(events, args.batch_size)
    print(f"Submitting {len(events)} CVEs in {len(batches)} batch(es) of <= {args.batch_size}")

    if args.in_process:
        from apps.common.utils import load_yaml
        from apps.workflow_d.orchestrator import Orchestrator
        from apps.workflow_d.schemas import AnalyzeRequest

        cfg = load_yaml(ROOT / "configs" / "app.yaml")
        orch = Orchestrator(cfg)
        for i, b in enumerate(batches, 1):
            req = AnalyzeRequest(mode=args.mode, cves=b)
            resp = orch.analyze(req)
            print(f"  batch {i}/{len(batches)} -> analysis_id={resp.analysis_id}")
            for r in resp.results:
                print(
                    f"    {r.cve_id} :: {r.component.name}@{r.component.current_version} "
                    f"-> verdict={r.routing.final_verdict.value} "
                    f"decision={r.routing.decision.value} "
                    f"conf=({r.confidence.triage_confidence},"
                    f"{r.confidence.fix_confidence},"
                    f"{r.confidence.evidence_confidence})"
                )
        return

    import requests

    for i, b in enumerate(batches, 1):
        resp = requests.post(
            f"{args.api.rstrip('/')}/analyze",
            json={"mode": args.mode, "cves": b},
            timeout=600,
        )
        if resp.status_code != 200:
            sys.exit(f"batch {i} failed: HTTP {resp.status_code}: {resp.text[:400]}")
        data = resp.json()
        print(f"  batch {i}/{len(batches)} -> analysis_id={data['analysis_id']}")
        for r in data["results"]:
            print(
                f"    {r['cve_id']} :: {r['component']['name']}@{r['component'].get('current_version')} "
                f"-> verdict={r['routing']['final_verdict']} "
                f"decision={r['routing']['decision']} "
                f"conf=({r['confidence']['triage_confidence']},"
                f"{r['confidence']['fix_confidence']},"
                f"{r['confidence']['evidence_confidence']})"
            )


if __name__ == "__main__":
    main()
